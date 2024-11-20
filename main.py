import psycopg2 as ps
from tabulate import tabulate
import openpyxl as xl
import art
import datetime

conn = ps.connect(
    host="localhost",
    dbname="postgres",
    user="app",
    password="1234",
    port=5432
    )

conn.autocommit = True
cursor = conn.cursor()
correctOrderOfTables = []

def createTeables():
    cursor.execute("DROP TABLE IF EXISTS Sales")
    cursor.execute("DROP TABLE IF EXISTS Orders")
    cursor.execute("DROP TABLE IF EXISTS Customers")
    cursor.execute("DROP TABLE IF EXISTS Providers")
    cursor.execute("DROP TABLE IF EXISTS Goods")

    cursor.execute("""
            CREATE TABLE Customers (
                Id SERIAL PRIMARY KEY,
                NPS TEXT,
                Phone TEXT,
                Email TEXT,
                Address TEXT
            );
        """)
    cursor.execute("""
            CREATE TABLE Providers (
                Id SERIAL PRIMARY KEY,
                Name TEXT,
                StartOfContract DATE,
                Email TEXT
            );
        """)
    cursor.execute("""
            CREATE TABLE Goods (
                Id SERIAL PRIMARY KEY,
                Name TEXT NOT NULL,
                Price INT NOT NULL,
                Count INT
            );
        """)
    cursor.execute("""
            CREATE TABLE Orders (
                Id SERIAL PRIMARY KEY,
                IdOfProduct INT NOT NULL,
                IdOfProvider INT NOT NULL,
                Date DATE,
                Count INT,
                FOREIGN KEY (IdOfProduct) REFERENCES Goods(Id),
                FOREIGN KEY (IdOfProvider) REFERENCES Providers(Id)
            );
        """)
    cursor.execute("""
            CREATE TABLE Sales (
                Id SERIAL PRIMARY KEY,
                IdOfProduct INT NOT NULL,
                IdOfCustomer INT NOT NULL,
                Date DATE,
                Count INT,
                FOREIGN KEY (IdOfProduct) REFERENCES Goods(Id),
                FOREIGN KEY (IdOfCustomer) REFERENCES Customers(Id)
            );
        """)

def matchSheets(sheet):
    match sheet:
        case "Клиенты":
            return "customers"
        case "Товары":
            return "goods"
        case "Поставщик":
            return "providers"
        case "Заказы":
            return "orders"
        case "Продажа":
            return "sales"
        case "customers":
            return "Клиенты"
        case "goods":
            return "Товары"
        case "providers":
            return "Поставщик"
        case "orders":
            return "Заказы"
        case "sales":
            return "Продажа"
        case _:
            return 0
        
def matchColumns(column):
    match column:
        case "id":
            return "№"
        case "nps":
            return "ФИО"
        case "phone":
            return "Номер тел."
        case "email":
            return "Почта"
        case "address":
            return "Адрес"
        case "name":
            return "Название"
        case "price":
            return "Цена"
        case "count":
            return "Кол-во"
        case "startofcontract":
            return "Дата заключения контракта"
        case "idofproduct":
            return "Номер товара"
        case "idofprovider":
            return "Номер поставщика"
        case "date":
            return "Дата"
        case "idofcustomer":
            return "Номер клиента"
        case _:
            return 0

def fromXlToPg():
    workbook = xl.load_workbook("example.xlsx")
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        nameOfSheet = matchSheets(sheet_name)

        if nameOfSheet == 0:
            continue

        cursor.execute(f"""SELECT column_name FROM information_schema.columns WHERE table_name = %s;""", (nameOfSheet,))
        columns = cursor.fetchall()
        correctOrderOfTables.append(nameOfSheet)

        column_names = [column[0] for column in columns]
        i = -1
        for row in sheet.iter_rows(values_only=True):
            if i < 0:
                i = i + 1
                continue
            placeholders = ', '.join(['%s'] * len(columns))
            noc = ', '.join(column_names)
            cursor.execute(f"""INSERT INTO {nameOfSheet} ({noc}) VALUES ({placeholders})""", row)

def getListOfColumns(pgPage, case):
    listOfColumns = []
    cursor.execute(f"""SELECT column_name FROM information_schema.columns WHERE table_name = %s""", (pgPage,))
    for column in cursor.fetchall():
        if int(case) == 1:
            listOfColumns.append(matchColumns(''.join(column)))
        else:
            listOfColumns.append(''.join(column))
    return listOfColumns

def getData(pgPage):
    cursor.execute(f"""SELECT * FROM {pgPage}""")
    data = cursor.fetchall()
    formattedList = []
    for item in data:
        newList = []
        for element in item:
            if isinstance(element, datetime.date):
                formatted_date = element.strftime("%d.%m.%Y")
                newList.append(formatted_date)
            else:
                newList.append(element)
        formattedList.append(list(newList))
    return formattedList

def showTable(pgPage):
    header = getListOfColumns(pgPage, 1)
    data = getData(pgPage)
    print("\n", matchSheets(pgPage), "\n", tabulate(data, headers=header, tablefmt="pretty"))

def action(pgPage):

    def checkItem(choise):
        cursor.execute(f"""SELECT COUNT(*) FROM {pgPage} WHERE id = %s;""", (choise,))
        res = cursor.fetchone()[0]
        return res
    
    def checkOfDate(date):
        try:
            return bool(datetime.datetime.strptime(date, "%d.%m.%Y"))
        except ValueError:
            print("Неверно введено значение")
            return False
        
    def createStrForChanging(newRow):
        prtOfReq = ""
        i = 1
        for column in getListOfColumns(pgPage, 2)[1:]:
            if str(newRow[i]).isdigit():
                if i == len(newRow) - 1:
                    prtOfReq = prtOfReq + str(column) + " = " + str(newRow[i])
                else:
                    prtOfReq = prtOfReq + str(column) + " = " + str(newRow[i]) + ", "
            else:
                if i == len(newRow) - 1:
                    prtOfReq = prtOfReq + str(column) + " = " + "\'" + str(newRow[i]) + "\'"
                else:
                    prtOfReq = prtOfReq + str(column) + " = " + "\'" + str(newRow[i]) + "\'" + ", "
            i = i + 1
        return prtOfReq
    
    def createStrForAdding(newRow):
        prtOfReq = ""
        i = 0
        for item in newRow:
            if str(item).isdigit():
                if i == len(newRow) - 1:
                    prtOfReq = prtOfReq + str(item)
                else:
                    prtOfReq = prtOfReq + str(item) + ", "
            else:
                if i == len(newRow) - 1:
                    prtOfReq = prtOfReq + "\'" + str(item) + "\'"
                else:
                    prtOfReq = prtOfReq + "\'" + str(item) + "\'" + ", "
            i = i + 1
        return prtOfReq
        
    def createNewRow(case):

        def checkOfCon(text, table):
            listOfRet = []
            itemList = getData(table)
            for i in itemList:
                sItemList = []
                for b in i[:2]:
                    sItemList.append(b)
                listOfRet.append(sItemList)
            for i in range(len(listOfRet)):
                print("Номер:", listOfRet[i][0], "\tНазвание:", listOfRet[i][1])
            print(text)
            bl = False
            while bl == False:
                prdct = input()
                for item in listOfRet:
                    if int(item[0]) == int(prdct):
                        bl = True
                        break
                if bl == False:
                    print("Неверно введено значение")
            return prdct

        changedRow = []

        if case == "2":
            while True:
                choise = int(input("Введите id строки: "))
                if checkItem(choise) == 0:
                    print("Такой строки не существует")
                else:
                    changedRow.append(choise)
                    break
        else:
            listOfId = []
            for i in getData(pgPage):
                listOfId.append(int(i[0]))
            num = 1
            while True:
                if listOfId.count(num):
                    num = num + 1
                else:
                    break
            changedRow.append(num)


        for col in getListOfColumns(pgPage, 1)[1:]:
            if col == "Дата" or col == "Дата заключения контракта":
                while True:
                    date = input("Введите дату формата DD.MM.YYYY (пример: 18.05.2001): ")
                    if checkOfDate(date) == True:
                        changedRow.append(date)
                        break
            elif col == "Номер товара":
                numbOfItem = checkOfCon("Введите Id товара из списка", "goods")
                changedRow.append(int(numbOfItem))
            elif col == "Номер клиента":
                numbOfItem = checkOfCon("Введите Id клиента из списка", "customers")
                changedRow.append(int(numbOfItem))
            elif col == "Номер поставщика":
                numbOfItem = checkOfCon("Введите Id поставщика из списка", "providers")
                changedRow.append(int(numbOfItem))
            else:
                changedRow.append(input("Введите " + str(col) + ": "))

        return changedRow
    

    showTable(pgPage)
    print("Выберите действие:\n1)Удалить строку\n2)Редактировать строку\n3)Добавить строку")
    choise = input()
    match choise:
        case "1":
            choise = int(input("Введите id строки: "))
            match pgPage:
                case "goods":
                    cursor.execute(f"""DELETE FROM sales WHERE idofproduct = %s""", (choise,))
                    cursor.execute(f"""DELETE FROM orders WHERE idofproduct = %s""", (choise,))
                    cursor.execute(f"""DELETE FROM goods WHERE id = %s""", (choise,))
                case "providers":
                    cursor.execute(f"""DELETE FROM orders WHERE idofprovider = %s""", (choise,))
                    cursor.execute(f"""DELETE FROM providers WHERE id = %s""", (choise,))
                case "customers":
                    cursor.execute(f"""DELETE FROM sales WHERE idofcustomer = %s""", (choise,))
                    cursor.execute(f"""DELETE FROM customers WHERE id = %s""", (choise,))
                case _:
                    cursor.execute(f"""DELETE FROM {pgPage} WHERE id = %s""", (choise,))

        case "2":
            newRow = createNewRow(choise)
            cursor.execute(f"""UPDATE {pgPage} SET {createStrForChanging(newRow)} WHERE Id = {int(newRow[0])}""")

        case "3":
            newRow = createNewRow(choise)
            if pgPage == "orders":
                cursor.execute(f"""SELECT count FROM goods WHERE id = {newRow[1]}""")
                summ = cursor.fetchone()[0] + int(newRow[len(newRow) - 1])
                cursor.execute(f"""UPDATE goods SET count = {summ} WHERE id = {newRow[1]}""")
            elif pgPage == "sales":
                cursor.execute(f"""SELECT count FROM goods WHERE id = {newRow[1]}""")
                count = cursor.fetchone()[0]
                while True:
                    summ = summ = count - int(newRow[len(newRow) - 1])
                    if summ < 0:
                        print(f"\nВы не можете продать товара больше чем есть на складе (сейчас на складе: {count}). Заполните форму заново.")
                        newRow = createNewRow(choise)
                    else:
                        cursor.execute(f"""UPDATE goods SET count = {summ} WHERE id = {newRow[1]}""")
                        break
            cursor.execute(f"""INSERT INTO {pgPage} VALUES ({createStrForAdding(newRow)})""")
                                
        case _:
            print("Неверно введено значение")

def getStatistics():
    cursor.execute("""SELECT 
                        p.id AS supplier_id,
                        p.Name AS supplier_name,
                        SUM(o.count) AS total_deliveries
                        FROM providers p
                        JOIN orders o ON p.id = o.idofprovider
                        GROUP BY p.id, p.Name
                        ORDER BY total_deliveries DESC
                        LIMIT 5;""")
    print("\nТоп 5 поставщиков:\n",tabulate(cursor.fetchall(), headers=["id", "Название", "Кол-во"], tablefmt="pretty"))
    cursor.execute("""SELECT 
                        t.id AS product_id,
                        t.Name AS product_name,
                        SUM(s.count) AS total_sold
                        FROM goods t
                        JOIN sales s ON t.id = s.idofproduct
                        GROUP BY t.id, t.Name
                        ORDER BY 
                        total_sold DESC 
                        LIMIT 5;""")
    print("\nТоп 5 самых продаваемых товаров:\n",tabulate(cursor.fetchall(), headers=["id", "Название", "Кол-во"], tablefmt="pretty"))
    cursor.execute("""SELECT COUNT(*) FROM customers""")
    print("\nКоличество покупателей: ", cursor.fetchone()[0])

def mainPage():
    art.tprint("Developed_by_Akreall")
    while True:
        for pgPage in correctOrderOfTables:
            showTable(pgPage)
        print("1)Редактирование таблицы\n2)Показать статистику\n3)Выход из программы")
        choise = input()
        match choise:
            case "1":
                while True:
                    print("Введите название таблицы")
                    nameOfTable = input()
                    match nameOfTable:
                        case "Клиенты":
                            action(matchSheets(nameOfTable))
                            break
                        case "Товары":
                            action(matchSheets(nameOfTable))
                            break
                        case "Поставщик":
                            action(matchSheets(nameOfTable))
                            break
                        case "Заказы":
                            action(matchSheets(nameOfTable))
                            break
                        case "Продажа":
                            action(matchSheets(nameOfTable))
                            break
                        case _:
                            print("Введено не верное значение")
                break
            case "2":
                getStatistics()
                break
            case "3":
                print("Завершение программы")
                break
            case _:
                print("Неверно введенное значение, повторите попытку")

def fromPgToXl():
    xlFile = 'example.xlsx'
    try:                                #Очистка excel файла
        wb = xl.load_workbook(xlFile)
        for sheet in correctOrderOfTables:
            std = wb[matchSheets(sheet)]
            wb.remove(std)  
    except FileNotFoundError:
        wb = xl.Workbook()

    for pgPage in correctOrderOfTables: #Добавление листов в файл
        
        listOfColumns = getListOfColumns(pgPage, 1)

        formattedList = getData(pgPage)
        
        sheetTitle = matchSheets(pgPage)
        sheet = wb.create_sheet(title=sheetTitle)
        sheet.append(listOfColumns)
        for row in formattedList:
            sheet.append(row)

    if "Sheet" in wb.sheetnames:
        std = wb["Sheet"]
        wb.remove(std)
    wb.save(xlFile)

def main():
    createTeables()
    fromXlToPg()
    mainPage()
    fromPgToXl()
    conn.close()

main()